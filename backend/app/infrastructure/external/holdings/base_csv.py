from __future__ import annotations

import csv
import json
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any, Iterable

import httpx
import xlrd
from openpyxl import load_workbook

from app.config import get_settings
from app.infrastructure.external.base import with_backoff


class CsvHoldingsProviderBase:
    async def download_bytes(self, url: str) -> bytes:
        settings = get_settings()

        async def operation() -> bytes:
            async with httpx.AsyncClient(
                timeout=settings.holdings_http_timeout,
                headers={"User-Agent": "quantbot/0.1 holdings tracker"},
            ) as client:
                response = await client.get(url)
                response.raise_for_status()
                return response.content

        return await with_backoff(operation)

    async def download_csv(self, url: str) -> list[dict[str, str]]:
        return self.parse_csv(self._decode_text(await self.download_bytes(url)))

    async def download_xlsx(
        self,
        url: str,
        *,
        sheet: str | None = None,
        header_contains: str | Iterable[str] | None = None,
    ) -> list[dict[str, Any]]:
        return self.parse_xlsx(
            await self.download_bytes(url),
            sheet=sheet,
            header_contains=header_contains,
        )

    async def download_xls(
        self,
        url: str,
        *,
        sheet: str | None = None,
        header_contains: str | Iterable[str] | None = None,
    ) -> list[dict[str, Any]]:
        return self.parse_xls(
            await self.download_bytes(url),
            sheet=sheet,
            header_contains=header_contains,
        )

    async def download_json(self, url: str) -> Any:
        return json.loads(self._decode_text(await self.download_bytes(url)))

    def parse_csv(self, text: str) -> list[dict[str, str]]:
        return [
            {self.normalize_header(key): value for key, value in row.items() if key is not None}
            for row in csv.DictReader(StringIO(text))
        ]

    def parse_csv_with_preamble(
        self,
        text: str,
        *,
        header_contains: str | Iterable[str],
    ) -> tuple[dict[str, str], list[dict[str, str]]]:
        raw_rows = list(csv.reader(StringIO(text)))
        header_index = self._find_header_index(raw_rows, header_contains)
        metadata = self._metadata_from_rows(raw_rows[:header_index])
        rows = self._dict_rows(raw_rows[header_index], raw_rows[header_index + 1 :])
        return metadata, rows

    def parse_xlsx(
        self,
        data: bytes,
        *,
        sheet: str | None = None,
        header_contains: str | Iterable[str] | None = None,
    ) -> list[dict[str, Any]]:
        _, rows = self.parse_xlsx_with_preamble(
            data,
            sheet=sheet,
            header_contains=header_contains,
        )
        return rows

    def parse_xlsx_with_preamble(
        self,
        data: bytes,
        *,
        sheet: str | None = None,
        header_contains: str | Iterable[str] | None = None,
    ) -> tuple[dict[str, str], list[dict[str, Any]]]:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        worksheet = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
        raw_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        target = header_contains or self._first_non_empty_row(raw_rows)
        header_index = self._find_header_index(raw_rows, target)
        metadata = self._metadata_from_rows(raw_rows[:header_index])
        rows = self._dict_rows(raw_rows[header_index], raw_rows[header_index + 1 :])
        workbook.close()
        return metadata, rows

    def parse_xls(
        self,
        data: bytes,
        *,
        sheet: str | None = None,
        header_contains: str | Iterable[str] | None = None,
    ) -> list[dict[str, Any]]:
        _, rows = self.parse_xls_with_preamble(
            data,
            sheet=sheet,
            header_contains=header_contains,
        )
        return rows

    def parse_xls_with_preamble(
        self,
        data: bytes,
        *,
        sheet: str | None = None,
        header_contains: str | Iterable[str] | None = None,
    ) -> tuple[dict[str, str], list[dict[str, Any]]]:
        workbook = xlrd.open_workbook(file_contents=data)
        worksheet = workbook.sheet_by_name(sheet) if sheet else workbook.sheet_by_index(0)
        raw_rows = [
            [
                self._xls_cell_value(workbook, worksheet.cell(row, column))
                for column in range(worksheet.ncols)
            ]
            for row in range(worksheet.nrows)
        ]
        target = header_contains or self._first_non_empty_row(raw_rows)
        header_index = self._find_header_index(raw_rows, target)
        metadata = self._metadata_from_rows(raw_rows[:header_index])
        rows = self._dict_rows(raw_rows[header_index], raw_rows[header_index + 1 :])
        return metadata, rows

    def normalize_header(self, value: str) -> str:
        return value.strip().lower()

    def parse_date(self, value: Any) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        clean = str(value).strip()
        lower = clean.lower()
        if lower.startswith("as of "):
            clean = clean[6:].strip()
        elif " as of " in lower:
            clean = clean[lower.rfind(" as of ") + len(" as of ") :].strip()
        if "T" in clean:
            try:
                return datetime.fromisoformat(clean.replace("Z", "+00:00")).date()
            except ValueError:
                pass
        for pattern in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%b %d, %Y", "%d-%b-%Y", "%Y%m%d"):
            try:
                return datetime.strptime(clean, pattern).date()
            except ValueError:
                continue
        raise ValueError(f"Unsupported date value: {value}")

    def parse_number(self, value: Any) -> float | None:
        if value is None:
            return None
        if isinstance(value, int | float):
            return float(value)
        clean = str(value).strip()
        clean = clean.replace("\u2212", "-").replace("\u2013", "-").replace("\u2014", "-")
        if not clean or clean.upper() in {
            "--",
            "-",
            "N/A",
            "NA",
            "NAN",
            "NULL",
            "NONE",
            "USD CASH",
        }:
            return None
        clean = clean.replace("$", "").replace(",", "").replace("%", "")
        clean = clean.replace("(", "-").replace(")", "")
        try:
            return float(clean)
        except ValueError:
            return None

    def clean_holding_ticker(self, value: Any) -> str | None:
        if value is None:
            return None
        clean = str(value).strip().upper()
        if clean in {
            "",
            "--",
            "-",
            "N/A",
            "NA",
            "NAN",
            "CASH",
            "USD",
            "USD CASH",
            "CASH_USD",
            "US DOLLAR",
            "US DOLLARS",
        }:
            return None
        return clean

    def _decode_text(self, data: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                return data.decode(encoding)
            except UnicodeDecodeError:
                continue
        return data.decode("utf-8", errors="replace")

    def _find_header_index(
        self,
        rows: list[list[Any]],
        header_contains: str | Iterable[str],
    ) -> int:
        needles = self._headers_to_find(header_contains)
        for index, row in enumerate(rows):
            normalized = {self.normalize_header(str(cell)) for cell in row if cell is not None}
            if needles <= normalized:
                return index
        raise ValueError(f"Could not find holdings header containing: {sorted(needles)}")

    def _headers_to_find(self, header_contains: str | Iterable[str]) -> set[str]:
        if isinstance(header_contains, str):
            return {self.normalize_header(header_contains)}
        return {self.normalize_header(str(value)) for value in header_contains}

    def _first_non_empty_row(self, rows: list[list[Any]]) -> str:
        for row in rows:
            for cell in row:
                if cell not in (None, ""):
                    return str(cell)
        raise ValueError("No non-empty rows found")

    def _dict_rows(self, header: list[Any], rows: list[list[Any]]) -> list[dict[str, Any]]:
        keys = [self.normalize_header(str(value)) if value is not None else "" for value in header]
        parsed: list[dict[str, Any]] = []
        for row in rows:
            if not any(cell not in (None, "") for cell in row):
                continue
            parsed.append({key: value for key, value in zip(keys, row, strict=False) if key})
        return parsed

    def _metadata_from_rows(self, rows: list[list[Any]]) -> dict[str, str]:
        metadata: dict[str, str] = {}
        for row in rows:
            cells = [str(cell).strip() for cell in row if cell not in (None, "")]
            for cell in cells:
                lower = cell.lower()
                if lower.startswith("as of ") or " as of " in lower:
                    metadata.setdefault("as of", cell)
            if len(cells) >= 2:
                metadata[self.normalize_header(cells[0].rstrip(":"))] = cells[1]
            elif len(cells) == 1 and ":" in cells[0]:
                key, value = cells[0].split(":", 1)
                metadata[self.normalize_header(key)] = value.strip()
        return metadata

    def _xls_cell_value(self, workbook: xlrd.Book, cell: xlrd.sheet.Cell) -> Any:
        if cell.ctype == xlrd.XL_CELL_EMPTY:
            return None
        if cell.ctype == xlrd.XL_CELL_DATE:
            return xlrd.xldate.xldate_as_datetime(cell.value, workbook.datemode)
        return cell.value
